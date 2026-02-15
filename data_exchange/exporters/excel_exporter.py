"""
Excel export functionality using openpyxl.
"""
import io
from typing import List, Dict, Any, Callable
from datetime import datetime, date
from django.db.models import QuerySet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Generic Excel exporter for Django models.
    """

    # Style constants
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DATA_ALIGNMENT = Alignment(vertical='center')
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, queryset: QuerySet, fields: List[str] = None,
                 field_labels: Dict[str, str] = None,
                 field_formatters: Dict[str, Callable] = None,
                 date_format: str = '%Y-%m-%d',
                 datetime_format: str = '%Y-%m-%d %H:%M:%S',
                 sheet_name: str = 'Data'):
        self.queryset = queryset
        self.model = queryset.model
        self.fields = fields or self._get_default_fields()
        self.field_labels = field_labels or {}
        self.field_formatters = field_formatters or {}
        self.date_format = date_format
        self.datetime_format = datetime_format
        self.sheet_name = sheet_name

    def _get_default_fields(self) -> List[str]:
        return [f.name for f in self.model._meta.fields]

    def _get_header(self) -> List[str]:
        headers = []
        for field in self.fields:
            label = self.field_labels.get(field, field.replace('_', ' ').title())
            headers.append(label)
        return headers

    def _get_value(self, obj: Any, field_name: str) -> Any:
        value = obj
        for part in field_name.split('__'):
            if value is None:
                break
            value = getattr(value, part, None)

        if field_name in self.field_formatters:
            return self.field_formatters[field_name](value)

        if value is None:
            return ''

        if callable(value):
            return value()

        if hasattr(value, 'pk'):
            return str(value)

        return value

    def _apply_header_style(self, ws, row: int, col: int):
        cell = ws.cell(row=row, column=col)
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = self.HEADER_ALIGNMENT
        cell.border = self.THIN_BORDER

    def _apply_data_style(self, ws, row: int, col: int):
        cell = ws.cell(row=row, column=col)
        cell.alignment = self.DATA_ALIGNMENT
        cell.border = self.THIN_BORDER

    def _auto_size_columns(self, ws):
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_workbook(self) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        headers = self._get_header()
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            self._apply_header_style(ws, 1, col)

        for row_idx, obj in enumerate(self.queryset, 2):
            for col_idx, field in enumerate(self.fields, 1):
                value = self._get_value(obj, field)

                if isinstance(value, datetime):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    ws.cell(row=row_idx, column=col_idx).number_format = 'YYYY-MM-DD HH:MM:SS'
                elif isinstance(value, date):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    ws.cell(row=row_idx, column=col_idx).number_format = 'YYYY-MM-DD'
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)

                self._apply_data_style(ws, row_idx, col_idx)

        self._auto_size_columns(ws)
        ws.freeze_panes = 'A2'

        return wb

    def export_to_bytes(self) -> bytes:
        wb = self.export_to_workbook()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_to_file(self, file_path: str):
        wb = self.export_to_workbook()
        wb.save(file_path)

    def export_to_response(self, filename: str = 'export.xlsx'):
        from django.http import HttpResponse

        content = self.export_to_bytes()

        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
